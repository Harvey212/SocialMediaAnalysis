from openpyxl import Workbook
import os
import glob
import openpyxl
import numpy as np
#############################33
from sklearn.metrics import accuracy_score
from sklearn.neural_network import MLPClassifier

from sklearn.metrics import recall_score
from sklearn import metrics
from sklearn.metrics import roc_auc_score
###################################################
import torch
import torch.nn as nn
import torch.optim as optim

###################################################
#er=['1','2','3']
#for y in range(len(er)-1,-1,-1):
#	print(er[y])

normalize=10000000
cat=['travel','unbox','challenge','culture','game','personal','talk','education','music']
###############################
cuttingline = '2021-12-18'
thisyear=cuttingline[0:4]
#####################################
videoTypeTotalView = dict()
videoTypeTotalCount = dict()
############################
par=".\\data"
folders = os.listdir(par)
for folder in folders:
	chil=os.path.join(par,folder,'*.xlsx')
	myid = str(folder)
	#print(folder)
	#print(myid)
	#################################
	videoTypeTotalView[myid] = {}
	videoTypeTotalCount[myid] = {}
	##################################
	files=glob.glob(chil)
	for k in range(len(files)):
		path = files[k]

		wb_obj = openpyxl.load_workbook(path)
		sheet_obj = wb_obj.active
		#####################
		#for m in range(2,sheet_obj.max_row+1):
		#	feat=(sheet_obj.cell(row = m, column = 7).value)
		#	date=(sheet_obj.cell(row = m, column = 4).value)
		#	print(date)
		####################################################
		for m in range(sheet_obj.max_row,1,-1):
			date=(sheet_obj.cell(row = m, column = 4).value)
			view=(sheet_obj.cell(row = m, column = 5).value)
			topic=(sheet_obj.cell(row = m, column = 6).value)
			feat=(sheet_obj.cell(row = m, column = 7).value)
			year = date[0:4]

			if topic=='persnoal':
				print(path)
				print(m)

			tempdic1=videoTypeTotalView[myid]
			Tempdic1=videoTypeTotalCount[myid]
			################################################3
			if year not in tempdic1.keys():
				tempdic1[year] ={}
				Tempdic1[year]={}

				tempdic2=tempdic1[year]
				Tempdic2=Tempdic1[year]

				tempdic2[topic] = view
				Tempdic2[topic] =1

				tempdic1[year] =tempdic2
				Tempdic1[year]=Tempdic2
			else:
				tempdic2=tempdic1[year]
				Tempdic2=Tempdic1[year]

				if topic not in tempdic2.keys():
					tempdic2[topic] = view
					Tempdic2[topic] =1
				else:
					tempdic2[topic] +=view
					Tempdic2[topic] +=1

				tempdic1[year] =tempdic2
				Tempdic1[year]=Tempdic2
			########################################3
			videoTypeTotalView[myid] = tempdic1
			videoTypeTotalCount[myid] = Tempdic1
############################################################3
videoTypeAvgWhole = dict() #only to before cuttingline, by cat only
videoTypeAvg = dict() #to date, by year->cat
youtuberC=0
for ids in videoTypeTotalView.keys():

	####################################
	tem={}
	tem2={}
	tem3={}
	for ca in cat:
		tem[ca]=0
		tem2[ca]=0
		tem3[ca]=0
	videoTypeAvgWhole[ids]=tem
	######################################
	####################################################
	youtuberC+=1
	########################################
	ys=videoTypeTotalView[ids]
	ys2=videoTypeTotalCount[ids]
	###############################
	videoTypeAvg[ids] ={}
	ys3=videoTypeAvg[ids]
	for y in ys.keys():
		tops=ys[y]
		tops2=ys2[y]
		#########
		ys3[y]={}
		tops3=ys3[y]
		for t in tops.keys():
			tops3[t] = int(tops[t]/tops2[t])
			##############################
			if y<thisyear:
				#print(ids)
				tem2[t]+=tops[t]
				tem3[t]+=tops2[t]
			#############################
		ys3[y]=tops3
	#####################
	videoTypeAvg[ids]=ys3
	##############################3
	tem=videoTypeAvgWhole[ids]
	for t in tem.keys():
		if tem3[t]!=0:
			tem[t]=tem2[t]/tem3[t]
	videoTypeAvgWhole[ids]=tem
	#####################################

####################################






##############################################

catdict={}
for c in range(len(cat)):
	catdict[cat[c]]=c
###################################################3
whodict={}
numTowho={}

wo=0
for woo in videoTypeTotalView.keys():
	whodict[woo]=wo
	numTowho[wo] = woo
	wo+=1
#################################################

par2="socialnode2.xlsx"
wb_obj2 = openpyxl.load_workbook(par2)
sheet_obj2 = wb_obj2.active

perL=['摩羯座','水瓶座','雙魚座','牡羊座','金牛座','雙子座','巨蟹座','獅子座','處女座','天秤座','天蠍座','射手座']
pertonum ={}

for pp in range(len(perL)):
	pertonum[perL[pp]]=pp

sexL=['F','M','G']
sextonum={}
for ss in range(len(sexL)):
	sextonum[sexL[ss]]=ss

subcount ={}
sex ={}
personale={}
for m in range(2,sheet_obj2.max_row+1):
	skip=sheet_obj2.cell(row = m, column = 9).value

	if skip==0:
		iden=sheet_obj2.cell(row = m, column = 1).value
		sub=sheet_obj2.cell(row = m, column = 10).value
		sexx=sheet_obj2.cell(row = m, column = 11).value
		pers=sheet_obj2.cell(row = m, column = 12).value
		#################

		subcount[str(iden)]=int(sub)
		sex[str(iden)]=sexx
		personale[str(iden)]=pers
		#######################
##############################################



#################################################################
yerToyertype =np.zeros((youtuberC,youtuberC*len(cat))) #to cutting line only
yerToyertypeC=np.zeros((youtuberC,youtuberC*len(cat))) #to cutting line only

Recavg=np.zeros((youtuberC,youtuberC*len(cat)))#to cutting line only
Recavg2=np.zeros((youtuberC,youtuberC*len(cat)))#to cutting line only
##################################
#####################################################
par=".\\data"
folders = os.listdir(par)
for folder in folders:
	chil=os.path.join(par,folder,'*.xlsx')
	myid = str(folder)
	YY=videoTypeAvg[myid]

	files=glob.glob(chil)
	for k in range(len(files)):
		path = files[k]

		wb_obj = openpyxl.load_workbook(path)
		sheet_obj = wb_obj.active

		for m in range(sheet_obj.max_row,1,-1):
			date=(sheet_obj.cell(row = m, column = 4).value)
			view=(sheet_obj.cell(row = m, column = 5).value)
			topic=(sheet_obj.cell(row = m, column = 6).value)
			feat=(sheet_obj.cell(row = m, column = 7).value)
			year = date[0:4]

			if date<cuttingline:
				if feat !='-1':
					ff=feat.split(',')
					for w in range(len(ff)):
						wh=ff[w]

						if wh in whodict.keys():
							col=whodict[wh]*len(cat)+catdict[topic]
							######################################
							yy=YY[year]
							val = view/yy[topic]
							#####################################3
							yerToyertype[whodict[myid]][col]+=val
							yerToyertypeC[whodict[myid]][col]+=1


	###########################################33
	
#######################################################
for i in range(Recavg.shape[0]):
	fromid=numTowho[i]
	fromsubcount=subcount[fromid]
	for j in range(Recavg.shape[1]):
		secCount=int(j/len(cat))
		toid=numTowho[secCount]
		tosubcount=subcount[toid]

		if yerToyertypeC[i][j]!=0:
			Recavg[i][j]=yerToyertype[i][j]/yerToyertypeC[i][j]
			Recavg2[i][j]=(yerToyertype[i][j]/yerToyertypeC[i][j])*(tosubcount/(fromsubcount+tosubcount))

#############################################################
#for model 2 ,model 4 
#=>Recavg2
###################################################3
#for model 3
#=>Recavg
##################################################
#########################################################3
#model 4
#sex,person,cat,sub,relationship
oldrelArr=np.zeros((youtuberC,youtuberC))
fakec=0
judge=0
for z1 in range(yerToyertypeC.shape[0]):
	for z2 in range(yerToyertypeC.shape[1]):
		fakec+=1
		secc=int(z2/len(cat))
		if yerToyertypeC[z1][z2]>0:
			judge=1
		if fakec==len(cat):
			if judge==1:
				oldrelArr[z1][secc]=1
				oldrelArr[secc][z1]=1
			judge=0
			fakec=0
#############################





##########################################################
featuredim=len(sexL)+len(perL)+len(cat)+1+len(list(whodict.keys()))
featuredic={}

for whh in whodict.keys():
	fea=np.zeros((1,featuredim))
	######################
	myst=0
	mysex=sex[whh]
	mysexnum=sextonum[mysex]
	col=myst+mysexnum
	fea[0][col]=1
	#####################3
	myst=len(sexL)
	mypers=personale[whh]
	mypersnum=pertonum[mypers]
	col=myst+mypersnum
	fea[0][col]=1
	#########################
	myst=len(sexL)+len(perL)
	mycat=videoTypeAvgWhole[whh]
	for cc in range(len(cat)):
		col=myst+cc
		fea[0][col]=mycat[cat[cc]]/normalize
	###############################
	myst=len(sexL)+len(perL)+len(cat)
	mysub=subcount[whh]
	col=myst
	fea[0][col]=mysub/normalize
	####################################
	myst=len(sexL)+len(perL)+len(cat)+1
	wn=whodict[whh]
	for rr in range(youtuberC):
		if oldrelArr[wn][rr]==1:
			col=myst+rr
			fea[0][col]=1
	#################################
	featuredic[whh]=fea
##########################################

trainX=[]
trainY=[]

#only train from the combination that once has the relationship
for from1 in range(oldrelArr.shape[0]):
	for to1 in range(from1+1,oldrelArr.shape[1]):
		#########################################
		if oldrelArr[from1][to1]==1:
			#####################################
			for cc1 in range(len(cat)):
				typevec=np.zeros((1,len(cat)))
				typevec[0][cc1]=1
				######################
				rec=0
				if (Recavg2[from1][len(cat)*to1+cc1]!=0) and (Recavg2[to1][len(cat)*from1+cc1]!=0): 
					rec=(Recavg2[from1][len(cat)*to1+cc1]+Recavg2[to1][len(cat)*from1+cc1])
				else:
					if (Recavg2[from1][len(cat)*to1+cc1]==0) and (Recavg2[to1][len(cat)*from1+cc1]==0):
						rec=0
					else:
						if (Recavg2[from1][len(cat)*to1+cc1]!=0):
							rec=Recavg[from1][len(cat)*to1+cc1]
						if (Recavg2[to1][len(cat)*from1+cc1]!=0):
							rec=Recavg[to1][len(cat)*from1+cc1]
				#print(rec)

				y_truth=0
				if rec>=1:
					y_truth=1

				fromvec=featuredic[numTowho[from1]]
				tovec=featuredic[numTowho[to1]]
				###########################
				f1=(fromvec.tolist())[0] + (tovec.tolist())[0]
				F1=f1+ (typevec.tolist())[0]
				############################
				f2=(tovec.tolist())[0] + (fromvec.tolist())[0]
				F2=f2+ (typevec.tolist())[0]
				###############################
				trainX.append(F1)
				trainX.append(F2)
				trainY.append(y_truth)
				trainY.append(y_truth)
				#############################
				
##################################
#clf = MLPClassifier(solver='lbfgs', alpha=1e-5,
#                    hidden_layer_sizes=(30, 4), random_state=1,max_iter=1000)

#clf.fit(trainX, trainY)
##########################################
class BinaryClassificationModel(nn.Module):
    def __init__(self):
        super(BinaryClassificationModel, self).__init__()
        self.fc1 = nn.Linear(2*featuredim+len(cat), 50)
        self.fc2 = nn.Linear(50, 20)
        self.fc3 = nn.Linear(20, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = self.fc3(x)
        x = self.sigmoid(x)
        return x



weight0=0
weight1=0
for i in range(len(trainY)):
	if trainY[i]==1:
		weight1+=1

weight0=len(trainY)-weight1

#print(weight1)
#print(weight0)
#class_counts = torch.tensor([weight0, weight1])
#class_weights = torch.tensor([weight1/(weight1+weight0), weight0/(weight1+weight0)])


#class_weights = 1. / class_counts.float()
pos_weight= torch.tensor([15])
criterion = nn.BCEWithLogitsLoss(pos_weight=pos_weight)
#criterion = nn.CrossEntropyLoss(weight=class_weights)



model = BinaryClassificationModel()
#criterion = nn.BCELoss()
optimizer = optim.Adam(model.parameters(), lr=0.001)


num_epochs = 200
for epoch in range(num_epochs):
	model.train()


	trainX = torch.from_numpy(np.array(trainX)).type(torch.float32)
	trainY = torch.from_numpy(np.array(trainY)).type(torch.float32)
	###############################
	optimizer.zero_grad()
	outputs = model(trainX).squeeze()
	loss = criterion(outputs, trainY)
	loss.backward()
	optimizer.step()

	# Print the loss for every epoch
	print(f'Epoch {epoch+1}/{num_epochs}, Loss: {loss.item()}')

###################################################











testcomb=[]
for i in range(oldrelArr.shape[0]):
	for j in range(i+1,oldrelArr.shape[1]):
		if oldrelArr[i][j]==0:
			com=set()
			com.add(numTowho[i])
			com.add(numTowho[j])
			testcomb.append(com)
			#strr='('+str(i)+','+str(j)+')'
			#print(strr)
##################################
sofarno=[]
beginyes=[]
beginyes2={} #for topic
############################
par=".\\data"
folders = os.listdir(par)
for folder in folders:
	chil=os.path.join(par,folder,'*.xlsx')
	myid = str(folder)
	files=glob.glob(chil)
	for k in range(len(files)):
		path = files[k]
		wb_obj = openpyxl.load_workbook(path)
		sheet_obj = wb_obj.active
		#####################
		####################################################
		for m in range(sheet_obj.max_row,1,-1):
			date=(sheet_obj.cell(row = m, column = 4).value)
			view=(sheet_obj.cell(row = m, column = 5).value)
			topic=(sheet_obj.cell(row = m, column = 6).value)
			feat=(sheet_obj.cell(row = m, column = 7).value)
			if (date>=cuttingline) and (feat !='-1'):
				ff=feat.split(',')
				for w in range(len(ff)):
					wh=ff[w]
					if wh in whodict.keys():
						check=set()
						check.add(myid)
						check.add(wh)
						############33
						####################3
						if (check in testcomb):
							if (check not in beginyes):
								beginyes.append(check)							
							####################################
							if myid not in beginyes2.keys():
								ter1={}
								ter1[topic]=1

								ter2={}
								ter2[wh]=ter1

								beginyes2[myid]=ter2
							else:
								ter2= beginyes2[myid]

								if wh in ter2.keys():
									ter1=ter2[wh]
									ter1[topic]=1
									ter2[wh]=ter1
									beginyes2[myid]=ter2
								else:
									ter1={}
									ter1[topic]=1
									ter2[wh]=ter1
									beginyes2[myid]=ter2
							################################3
							if wh not in beginyes2.keys():
								ter1={}
								ter1[topic]=1

								ter2={}
								ter2[myid]=ter1

								beginyes2[wh]=ter2
							else:
								ter2= beginyes2[wh]

								if myid in ter2.keys():
									ter1=ter2[myid]
									ter1[topic]=1
									ter2[myid]=ter1
									beginyes2[wh]=ter2
								else:
									ter1={}
									ter1[topic]=1
									ter2[myid]=ter1
									beginyes2[wh]=ter2
							
##################################################
for see in testcomb:
	if see not in beginyes:
		sofarno.append(see)
#####################################
print(len(testcomb))
print(len(beginyes))
print(len(sofarno))
print('haaha')
###########################
#category
testx=[]
gtruthcat=[]
######################
#youtuber only
gtruthcat2=[]
##########################
for td in testcomb:
	td2=list(td)
	fromid=td2[0]
	toid=td2[1]
	fromvec=featuredic[fromid]
	tovec=featuredic[toid]
	#############################
	#########################
	gt2=0
	for cc1 in range(len(cat)):
		typevec=np.zeros((1,len(cat)))
		typevec[0][cc1]=1
		f1=(fromvec.tolist())[0] + (tovec.tolist())[0]
		F1=f1+ (typevec.tolist())[0]

		testx.append(F1)
		###########################################
		gt=0
		if (fromid in beginyes2.keys()):
			ter2=beginyes2[fromid]
			if toid in ter2.keys():
				ter1=ter2[toid]
				if cat[cc1] in ter1.keys():
					gt=1
					gt2=1

		gtruthcat.append(gt)
		########################################
	#####################################################
	gtruthcat2.append(gt2)


#########################
#include category
#y_pred1=clf.predict(testx)
#gtruthcat

model.eval()
with torch.no_grad():
	outputs = model(torch.tensor(np.array(testx), dtype=torch.float32)).squeeze()
	y_pred1 = (outputs > 0.5).float()
	#accuracy = (predictions == y_pred1).float().mean()
	#print(f'Accuracy: {accuracy.item() * 100:.2f}%')

###############################################################





################################
y_pred2=[]
#gtruthcat2

fakec=0
judge=0
for i in range(len(y_pred1)):
	fakec+=1
	if y_pred1[i]==1:
		judge=1

	if fakec==len(cat):
		if judge==1:
			y_pred2.append(1)
		else:
			y_pred2.append(0)

		fakec=0
		judge=0
#################################

seeee=y_pred1.tolist()
for i in range(len(seeee)):
	print(seeee[i])
#print(len(gtruthcat))

#print(len(y_pred2))
#print(len(gtruthcat2))
#################################
ac1=accuracy_score(gtruthcat, y_pred1)
ac2=accuracy_score(gtruthcat2, y_pred2)

print('accuracy for category')
print(ac1)

print('auc for category')

print(roc_auc_score(np.array(gtruthcat), np.array(y_pred1)))

print('f1 score for category')
print(metrics.f1_score(gtruthcat, y_pred1, average='weighted'))
##############################################
print('accuracy for youtuber')
print(ac2)

print('auc for youtuber')
print(roc_auc_score(np.array(gtruthcat2), np.array(y_pred2)))

print('f1 score for youtuber')
print(metrics.f1_score(gtruthcat2, y_pred2, average='weighted'))
###################################
forrecallgt=[]
forrecallpred=[]
y_pred1=y_pred1.tolist()
for i in range(len(gtruthcat)):
	if gtruthcat[i]==1:
		forrecallgt.append(1)

		forrecallpred.append(y_pred1[i])

print('Recall for category')
print(sum(forrecallpred)/sum(forrecallgt))
#print(recall_score(gtruthcat, y_pred1, average='macro'))

#########################
forrecallgt2=[]
forrecallpred2=[]

for i in range(len(gtruthcat2)):
	if gtruthcat2[i]==1:
		forrecallgt2.append(1)
		forrecallpred2.append(y_pred2[i])

print('Recall for youtuber')
print(sum(forrecallpred2)/sum(forrecallgt2))
#print(recall_score(gtruthcat2, y_pred2, average='macro'))
